#! CONVERTING BETWEEN COLUMN LETTERS AND NUMBERS
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

print(get_column_letter(1))
print(get_column_letter(2))
print(get_column_letter(27))
print(get_column_letter(900))

wb = openpyxl.load_workbook('D:\\auto_py\\chp12_excel\\example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
print(get_column_letter(sheet.max_column))
print(column_index_from_string('A'))
print(column_index_from_string('AA'))