import openpyxl

wb = openpyxl.load_workbook('D:\\auto_py\\chp12_excel\\example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
print(sheet.max_row)
print(sheet.max_column)