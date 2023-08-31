#! python3\
#readCencsusExcel.py - Tabulates population and number of census tracts for each country

import openpyxl, pprint
print('Opening Workbook.....')

wb = openpyxl.load_workbook('D:\\auto_py\\chp12_excel\\censuspopdata.xlsx')
sheet = wb.get_sheet_by_name('Population by Census Tract')
countyData = {}

# TODO: Fill in countyData with each county's population and tracts.
print('Reading rows...')
for row in range(2, sheet.max_row + 1):
    # Each row in the spreadsheet has data for one census tract.
    state = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    pop = sheet['D' + str(row)].value

# TODO: Open a new text file and write the contents of countyData to it.

for row in range(2, sheet.max_row + 1):
    # Each row in teh spreadsheet has data for one census tract.
    state = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    pop = sheet['D' + str(row)].value
    
    