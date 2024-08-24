#! python3
#updateProduce.py - Corrects costs in produce sales spreadsheet.

import openpyxl

wb = openpyxl.load_workbook('D:\\auto_py\\chp12_excel\\produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

#print(sheet.cell(row=1, column=1).value) #validate to make sure we are in the right file.
#THE PRODUCE TYPES AND THEIR UPDATED PRICES
PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27}

#TODO : LOOP THROUGH THE ROWS AND UPDATE THE PRICES

for rowNum in range(2, sheet.max_row): #skip the first row
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

if produceName == 'Celery':
    cellObj = 1.19
if produceName == 'Garlic':
    cellObj = 3.07
if produceName == 'Lemon':
    cellObj = 1.27

wb.save('chp12_excel\\updatedProduceSales.xlsx')